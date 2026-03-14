#!/usr/bin/env python3
"""Export all company data to a single .xlsx workbook."""

import os
import sys
from collections import defaultdict

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../modules/data/scripts"))
import datalib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SURFACE_ROOT = os.environ.get("SURFACE_ROOT", os.path.join(os.path.dirname(__file__), ".."))
DOWNLOADS_DIR = os.path.join(SURFACE_ROOT, "downloads")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")


def style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)


def add_cap_table(wb):
    share_data = datalib.load("shares")
    ct = datalib.cap_table(share_data)
    ws = wb.create_sheet("Cap Table")
    ws.append(["Holder", "Class", "Held", "%", "Vested"])
    for r in ct:
        ws.append([r["holder"], r["class"], r["held"], f"{r['pct']}%", r["held"]])
    ws.append(["Total", "", sum(r["held"] for r in ct), "", ""])
    style_header(ws)
    auto_width(ws)


def add_share_history(wb):
    share_data = datalib.load("shares")
    holders_map = {h["id"]: h["display_name"] for h in share_data.get("holders", [])}
    events = sorted(share_data.get("share_events", []), key=lambda e: str(e.get("event_date", "")))
    ws = wb.create_sheet("Share History")
    ws.append(["Date", "Event", "Holder", "Class", "Qty"])
    for e in events:
        ws.append([
            str(e.get("event_date", "")),
            e["event_type"],
            holders_map.get(e["holder_id"], e["holder_id"]),
            e["share_class"],
            e["quantity"],
        ])
    style_header(ws)
    auto_width(ws)


def add_holders(wb):
    share_data = datalib.load("shares")
    h = datalib.holdings(share_data)
    holdings_map = defaultdict(int)
    for r in h:
        holdings_map[r["holder_id"]] += r["shares_held"]
    ws = wb.create_sheet("Holders")
    ws.append(["ID", "Name", "Total Shares"])
    for holder in sorted(share_data.get("holders", []), key=lambda x: x["id"]):
        ws.append([holder["id"], holder["display_name"], holdings_map.get(holder["id"], 0)])
    style_header(ws)
    auto_width(ws)


def add_vesting(wb):
    share_data = datalib.load("shares")
    holders_map = {h["id"]: h["display_name"] for h in share_data.get("holders", [])}
    vs = datalib.vesting_schedule(share_data)
    ws = wb.create_sheet("Vesting")
    ws.append(["Holder", "Class", "Granted", "Vested", "Unvested", "% Vested", "Cliff", "Fully Vested"])
    for v in vs:
        ws.append([
            holders_map.get(v["holder_id"], v["holder_id"]),
            v["share_class"],
            v["total_granted"],
            v["vested"],
            v["unvested"],
            f"{v['pct_vested']}%",
            v["cliff_date"],
            v["fully_vested_date"],
        ])
    style_header(ws)
    auto_width(ws)


def add_pools(wb):
    share_data = datalib.load("shares")
    h = datalib.holdings(share_data)
    holders_map = {r["id"]: r["display_name"] for r in share_data.get("holders", [])}
    holdings_map = {}
    for r in h:
        holdings_map[(r["holder_id"], r["share_class"])] = r["shares_held"]
    ws = wb.create_sheet("Pools")
    ws.append(["Pool", "Class", "Budget", "Issued", "Available", "Members"])
    for pool in sorted(share_data.get("pools", []), key=lambda x: x["name"]):
        members = [pm for pm in share_data.get("pool_members", []) if pm["pool_name"] == pool["name"]]
        issued = sum(holdings_map.get((pm["holder_id"], pool["share_class"]), 0) for pm in members)
        member_strs = [
            f"{holders_map.get(pm['holder_id'], pm['holder_id'])} ({holdings_map.get((pm['holder_id'], pool['share_class']), 0)})"
            for pm in members
        ]
        ws.append([
            pool["name"],
            pool["share_class"],
            pool["budget"],
            issued,
            pool["budget"] - issued,
            ", ".join(member_strs) if member_strs else "-",
        ])
    style_header(ws)
    auto_width(ws)


def add_officers(wb):
    data = datalib.load("officers")
    ws = wb.create_sheet("Officers")
    ws.append(["ID", "Name", "Role", "Appointed", "Resigned"])
    for o in data.get("officers", []):
        ws.append([
            o["id"],
            o["person_name"],
            o["role"],
            str(o.get("appointed_date", "")),
            str(o.get("resigned_date", "")),
        ])
    style_header(ws)
    auto_width(ws)


def add_compliance(wb):
    data = datalib.load("compliance")
    ws = wb.create_sheet("Compliance")
    ws.append(["ID", "Title", "Due Date", "Frequency", "Category", "Status", "Filed"])
    for d in sorted(data.get("deadlines", []), key=lambda x: str(x.get("due_date", ""))):
        ws.append([
            d["id"],
            d["title"],
            str(d.get("due_date", "")),
            d.get("frequency", ""),
            d.get("category", ""),
            d.get("status", ""),
            str(d.get("filed_date", "")),
        ])
    style_header(ws)
    auto_width(ws)


def add_charges(wb):
    data = datalib.load("charges")
    charges = data.get("charges", [])
    if not charges:
        return
    ws = wb.create_sheet("Charges")
    ws.append(["ID", "Code", "Created", "Description", "Chargee", "Amount", "Currency", "Status"])
    for c in charges:
        ws.append([
            c["id"],
            c.get("charge_code", ""),
            str(c.get("created_date", "")),
            c["description"],
            c["chargee"],
            c.get("amount", 0),
            c.get("currency", "GBP"),
            c["status"],
        ])
    style_header(ws)
    auto_width(ws)


def add_dividends(wb):
    data = datalib.load("dividends")
    divs = data.get("dividends", [])
    if not divs:
        return
    ws = wb.create_sheet("Dividends")
    ws.append(["ID", "Class", "Per Share", "Currency", "Declared", "Paid", "Status"])
    for d in sorted(divs, key=lambda x: str(x.get("declaration_date", ""))):
        ws.append([
            d["id"],
            d.get("share_class", ""),
            d.get("amount_per_share", ""),
            d.get("currency", "GBP"),
            str(d.get("declaration_date", "")),
            str(d.get("payment_date", "")),
            d.get("status", ""),
        ])
    style_header(ws)
    auto_width(ws)


def add_company(wb):
    data = datalib.load("company")
    co = data.get("company", {})
    if not co:
        return
    ws = wb.create_sheet("Company")
    ws.append(["Field", "Value"])
    for key in ["name", "company_number", "jurisdiction", "company_type",
                "incorporation_date", "accounting_reference_date", "articles"]:
        if key in co:
            ws.append([key.replace("_", " ").title(), str(co[key])])
    addr = co.get("registered_address", {})
    if addr:
        addr_str = ", ".join(filter(None, [
            addr.get("line_1"), addr.get("line_2"),
            addr.get("city"), addr.get("postcode"), addr.get("country"),
        ]))
        ws.append(["Registered Address", addr_str])
    sic = co.get("sic_codes", [])
    if sic:
        ws.append(["SIC Codes", ", ".join(sic)])
    style_header(ws)
    auto_width(ws)


def main():
    os.makedirs(DOWNLOADS_DIR, exist_ok=True)
    output = os.path.join(DOWNLOADS_DIR, "formabi-data.xlsx")

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    add_company(wb)
    add_cap_table(wb)
    add_share_history(wb)
    add_holders(wb)
    add_vesting(wb)
    add_pools(wb)
    add_officers(wb)
    add_compliance(wb)
    add_charges(wb)
    add_dividends(wb)

    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
