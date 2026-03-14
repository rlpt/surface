#!/usr/bin/env python3
"""Export a standalone cap table spreadsheet."""

import os
import sys
from collections import defaultdict
from datetime import date

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../modules/data/scripts"))
import datalib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

SURFACE_ROOT = os.environ.get("SURFACE_ROOT", os.path.join(os.path.dirname(__file__), ".."))
DOWNLOADS_DIR = os.path.join(SURFACE_ROOT, "downloads")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
TOTAL_FONT = Font(bold=True)
PCT_FMT = '0.0"%"'


def style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)


def add_cap_table(wb, share_data):
    """Main cap table: holder, class, held, %, vested, unvested."""
    ct = datalib.cap_table(share_data)
    if not ct:
        return
    total = sum(r["held"] for r in ct)

    # Build vesting lookup
    vs = datalib.vesting_schedule(share_data)
    h = datalib.holdings(share_data)
    vested_lookup = defaultdict(int)
    for v in vs:
        vested_lookup[(v["holder_id"], v["share_class"])] += v["vested"]
    held_lookup = {}
    for r in h:
        held_lookup[(r["holder_id"], r["share_class"])] = r["shares_held"]
    for key in vested_lookup:
        if key in held_lookup:
            vested_lookup[key] = min(vested_lookup[key], held_lookup[key])

    ws = wb.create_sheet("Cap Table")
    ws.append(["Holder", "Class", "Held", "%", "Vested", "Unvested"])
    for r in ct:
        key = (r["holder_id"], r["class"])
        vested = vested_lookup.get(key, r["held"])
        ws.append([r["holder"], r["class"], r["held"], r["pct"], vested, r["held"] - vested])
    # Totals row
    row = ws.max_row + 1
    ws.append(["Total", "", total, 100.0, "", ""])
    for cell in ws[row]:
        cell.font = TOTAL_FONT

    # Format % column
    for row_cells in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row_cells:
            cell.number_format = PCT_FMT

    style_header(ws)
    auto_width(ws)


def add_share_classes(wb, share_data):
    """Share class summary with issued/available."""
    ca = datalib.class_availability(share_data)
    if not ca:
        return
    classes = {sc["name"]: sc for sc in share_data.get("share_classes", [])}

    ws = wb.create_sheet("Share Classes")
    ws.append(["Class", "Nominal Value", "Currency", "Authorised", "Issued", "Available", "% Issued"])
    for r in ca:
        sc = classes.get(r["class"], {})
        pct = round(r["issued"] * 100.0 / r["authorised"], 1) if r["authorised"] else 0
        ws.append([
            r["class"],
            sc.get("nominal_value", ""),
            sc.get("nominal_currency", ""),
            r["authorised"],
            r["issued"],
            r["available"],
            pct,
        ])
    for row_cells in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row_cells:
            cell.number_format = PCT_FMT
    style_header(ws)
    auto_width(ws)


def add_vesting(wb, share_data):
    """Vesting schedules — only if any grants have vesting."""
    vs = datalib.vesting_schedule(share_data)
    has_vesting = any(v["cliff_date"] or v["fully_vested_date"] for v in vs)
    if not has_vesting:
        return
    holders_map = {h["id"]: h["display_name"] for h in share_data.get("holders", [])}

    ws = wb.create_sheet("Vesting")
    ws.append(["Holder", "Class", "Granted", "Vested", "Unvested", "% Vested", "Cliff", "Fully Vested"])
    for v in vs:
        if not v["cliff_date"] and not v["fully_vested_date"]:
            continue
        ws.append([
            holders_map.get(v["holder_id"], v["holder_id"]),
            v["share_class"],
            v["total_granted"],
            v["vested"],
            v["unvested"],
            v["pct_vested"],
            v["cliff_date"],
            v["fully_vested_date"],
        ])
    for row_cells in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row_cells:
            cell.number_format = PCT_FMT
    style_header(ws)
    auto_width(ws)


def add_pools(wb, share_data):
    """Pool budgets and membership."""
    pools = share_data.get("pools", [])
    if not pools:
        return
    h = datalib.holdings(share_data)
    holders_map = {r["id"]: r["display_name"] for r in share_data.get("holders", [])}
    holdings_map = {}
    for r in h:
        holdings_map[(r["holder_id"], r["share_class"])] = r["shares_held"]

    ws = wb.create_sheet("Pools")
    ws.append(["Pool", "Class", "Budget", "Issued", "Available", "% Used", "Members"])
    for pool in sorted(pools, key=lambda x: x["name"]):
        members = [pm for pm in share_data.get("pool_members", []) if pm["pool_name"] == pool["name"]]
        issued = sum(holdings_map.get((pm["holder_id"], pool["share_class"]), 0) for pm in members)
        avail = pool["budget"] - issued
        pct = round(issued * 100.0 / pool["budget"], 1) if pool["budget"] else 0
        member_strs = [
            f"{holders_map.get(pm['holder_id'], pm['holder_id'])} ({holdings_map.get((pm['holder_id'], pool['share_class']), 0)})"
            for pm in members
        ]
        ws.append([
            pool["name"],
            pool["share_class"],
            pool["budget"],
            issued,
            avail,
            pct,
            ", ".join(member_strs) if member_strs else "-",
        ])
    for row_cells in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row_cells:
            cell.number_format = PCT_FMT
    style_header(ws)
    auto_width(ws)


def add_history(wb, share_data):
    """Full event ledger."""
    events = share_data.get("share_events", [])
    if not events:
        return
    holders_map = {h["id"]: h["display_name"] for h in share_data.get("holders", [])}

    ws = wb.create_sheet("Event History")
    ws.append(["Date", "Event", "Holder", "Class", "Qty"])
    for e in sorted(events, key=lambda e: str(e.get("event_date", ""))):
        ws.append([
            str(e.get("event_date", "")),
            e["event_type"],
            holders_map.get(e["holder_id"], e["holder_id"]),
            e["share_class"],
            e["quantity"],
        ])
    style_header(ws)
    auto_width(ws)


def main():
    os.makedirs(DOWNLOADS_DIR, exist_ok=True)
    output = os.path.join(DOWNLOADS_DIR, "cap-table.xlsx")

    share_data = datalib.load("shares")

    wb = Workbook()
    wb.remove(wb.active)

    add_cap_table(wb, share_data)
    add_share_classes(wb, share_data)
    add_vesting(wb, share_data)
    add_pools(wb, share_data)
    add_history(wb, share_data)

    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
